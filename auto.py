from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import traceback 
import openpyxl
import re
from openpyxl import load_workbook
from statistics import mean
from openpyxl.styles import Font, Alignment

lista_produtos = [] 
pesquisa = input("Digite um produto que queira buscar? ")

#Agente do navegador
options = webdriver.ChromeOptions() 
options.add_argument("--start-maximized") 
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

browser = webdriver.Chrome(options=options)
wait = WebDriverWait(browser, 10) #timesleep diferente

# AMAZON
try:
    print(f"\n--- Iniciando busca na AMAZON---")
    url_amazon = f"https://www.amazon.com.br/s?k={pesquisa}" #Salva URL em um variavel
    browser.get(url_amazon) #Abre URL
    
    cards = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-component-type="s-search-result"]')))    # Espera os cards carregarem
    cards_amazon = cards[:20]
    print(f"Amazon: Encontrei {len(cards)} produtos.")

    for card in cards_amazon:   #recolhe informações
        try:    #titulo
            titulo = card.find_element(By.CSS_SELECTOR, "div.s-main-slot h2").text   
        except:
            titulo = "Sem Título"
            
        try:    #preço
            p_int = card.find_element(By.CSS_SELECTOR, "span.a-price-whole").text
            p_cent = card.find_element(By.CSS_SELECTOR, "span.a-price-fraction").text
            preco = f"R$ {p_int},{p_cent}"  #concatenação, pois preço fracioado fica em divs diferentes
        except:
            preco = "Indisponível" 

        try:    #avaliação
            aval = card.find_element(By.CSS_SELECTOR, "div.a-row.a-size-small span").get_attribute("aria-label") 
            if not aval: 
                aval = card.find_element(By.CSS_SELECTOR, "div.a-row.a-size-small span").text
        except:
            aval = "Sem avaliação"
                
        try:   #link 
            link = card.find_element(By.CSS_SELECTOR, "div.s-product-image-container a.a-link-normal").get_attribute("href")
        except:
            link = "Sem link"

        lista_produtos.append({ #lista de produtos Amazon
            'loja': 'Amazon',
            'titulo': titulo,
            'preco': preco,
            'avaliacao': aval,
            'link': link
        })

except Exception:
    print("Erro na parte da Amazon (Pode continuar para Magalu):")
    traceback.print_exc()   #exibe o erro


try:
    print(f"\n--- Iniciando busca na MAGALU ---")
    url_magalu = f"https://www.magazineluiza.com.br/busca/{pesquisa}/"
    browser.get(url_magalu)
    
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="product-card-container"]')))
    cards = browser.find_elements(By.CSS_SELECTOR, '[data-testid="product-card-container"]')
    cards_magalu = cards[:20]
    print(f"Magalu: Encontrei {len(cards)} produtos.")

#-------------------------------------------------------------------------------------------
    for card in cards_magalu:
        texto_completo = card.text 
        if len(texto_completo) < 10:
            continue
        
        try:    #titulo
            titulo = card.find_element(By.CSS_SELECTOR, '[data-testid="product-title"]').text
        except:
            
            linhas = texto_completo.split('\n')
            titulo = linhas[0] if linhas else "Sem Título"

        preco_final_str = "Indisponível"
        try:    #preço (logica com ReGex)
            preco_elemento = card.find_element(By.CSS_SELECTOR, '[data-testid="price-value"]')
            preco_final_str = preco_elemento.text
        
        except:
            try:
                if "por R$" in texto_completo:
                    partes = texto_completo.split("por R$")
                    match = re.search(r'\s?[\d.,]+', partes[1])
                    if match:
                        preco_final_str = f"R${match.group()}"
                    else:
                        match = re.search(r'R\$\s?[\d.,]+', texto_completo)
                        if match:
                            preco_final_str = match.group()
            except:
                preco_final_str = "Erro Preço"
        if preco_final_str != "Indisponível" and preco_final_str != "Erro Preço":
            preco_final_str = preco_final_str.replace("ou ", "").strip()
            
        try:    #avaliação
            aval = card.find_element(By.CSS_SELECTOR, '[data-testid="review-primary-information"]').text
        except:
            match_nota = re.search(r"(\d[.,]\d)\s\((\d+)\)", texto_completo)
            if match_nota:
                aval = match_nota.group()
            else:
                match_qtd = re.search(r"\(\d+\)", texto_completo)
                if match_qtd:
                    aval = f"? {match_qtd.group()}" # retorna "? (120)"
                else:
                    aval = "Sem avaliação"

        try:    #link
            link = card.get_attribute("href")
            if not link:
                elemento_a = card.find_element(By.CSS_SELECTOR, "a")
                link = elemento_a.get_attribute("href")
        except:
            link = "Sem link"

        lista_produtos.append({
            'loja': 'Magalu',
            'titulo': titulo,
            'preco': preco_final_str,
            'avaliacao': aval,
            'link': link
        })

except Exception:
    print("Erro na parte da Magalu:")
    traceback.print_exc()

finally:
    browser.quit()  #fecha o navegador

print("\nRealizando cálculos matemáticos...")

def converter_para_float(preco_str):
    try:
        limpo = preco_str.replace("R$", "").strip()
        limpo = limpo.replace(".", "") # Tira ponto de milhar
        limpo = limpo.replace(",", ".") # Troca virgula por ponto
        return float(limpo)
    except:
        return 0.0

#-----------------------------------------------------------------------------------------
precos_amazon = []
precos_magalu = []
produtos_validos = []

for p in lista_produtos:
    valor = converter_para_float(p['preco'])
    
    if valor > 0:
        p['valor_num'] = valor
        produtos_validos.append(p)
        
        if p['loja'] == 'Amazon':
            precos_amazon.append(valor)
        elif p['loja'] == 'Magalu':
            precos_magalu.append(valor)


media_amazon = mean(precos_amazon) if precos_amazon else 0
media_magalu = mean(precos_magalu) if precos_magalu else 0

if produtos_validos:
    # Ordena a lista pelo valor numérico
    ordenados = sorted(produtos_validos, key=lambda x: x['valor_num'])
    produto_mais_barato = ordenados[0]
    produto_mais_caro = ordenados[-1]
else:
    produto_mais_barato = {'titulo': 'Nenhum', 'valor_num': 0}
    produto_mais_caro = {'titulo': 'Nenhum', 'valor_num': 0}


print("\nGerando planilhas...")

try:
    #   Lista completa
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Produtos"
    ws1.append(["Loja", "Produto", "Preço", "Avaliação", "Link"])

    for p in lista_produtos:
        ws1.append([p['loja'], p['titulo'], p['preco'], p['avaliacao'], p['link']])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    wb1.save("produtos_comparativo.xlsx")
    print("1. Arquivo 'produtos_comparativo.xlsx' criado.")

    #   Relatório estátistico
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "RelatorioGeral"
    
    ws2.append(["Resumo", "Nome do Produto", "Valor", "Link"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    #Mais barato
    ws2.append([
        "Produto Mais Barato", 
        produto_mais_barato['titulo'], 
        f"R$ {produto_mais_barato.get('valor_num', 0):.2f}", 
        produto_mais_barato.get('link', '-')
    ])

    #Mais caro
    ws2.append([
        "Produto Mais Caro", 
        produto_mais_caro['titulo'], 
        f"R$ {produto_mais_caro.get('valor_num', 0):.2f}", 
        produto_mais_caro.get('link', '-')
    ])

    # Medias
    ws2.append(["Média Amazon", "-", media_amazon, "-"])
    ws2.append(["Média Magalu", "-", media_magalu, "-"])

    wb2.save("relatorio_geral.xlsx")
    print("2. Arquivo 'relatorio_geral.xlsx' criado.")

except Exception:
    print("Erro ao salvar Excel:")
    traceback.print_exc()
    

#---------------------------------------------------
print("\nGerando planilhas dos exercícios...")

# Classificar produto por faixa de preço
valor_base = float(input("\nDigite um valor referência: "))     #Pede um valor de referência


acima_80 = []       #Listas para cada faixa de preço
entre_50_80 = []
abaixo_50 = []


for p in produtos_validos:        #Percorre os produtos que tem valor

    preco = p['valor_num']  #Preço
    print(f"Produto: R$ {preco} vs Limite 80%: R$ {valor_base*0.8}")

    if preco >= valor_base * 0.8:       #Acima de 80%
        acima_80.append(p)

    elif preco >= valor_base * 0.5:     #Entre 50% e 80%
        entre_50_80.append(p)
    
    else:
        abaixo_50.append(p)     #Abaixo de 50%


wbB = openpyxl.Workbook()       #Cria planilha
wsB = wbB.active
wsB.title = "faixabase"


wsB.append(["Faixa", "Loja", "Produto", "Preço", "Avaliação", "Link"])
for cell in wsB[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')      #Cabeçalho


for p in acima_80:      #Inseri produto acima de 80%
    wsB.append(["Acima da Faixa", p['loja'], p['titulo'], f"R${p['valor_num']}", p['avaliacao'], p['link']])


for p in entre_50_80:       #Inseri produto entre 50% e 80%
    wsB.append(["Dentro da Faixa", p['loja'], p['titulo'], f"R${p['valor_num']}", p['avaliacao'], p['link']])


for p in abaixo_50:     # Inseri produto abaixo de 50%
    wsB.append(["Muito barato", p['loja'], p['titulo'], f"R${p['valor_num']}", p['avaliacao'], p['link']])


wbB.save("valorbase.xlsx")        #Salva arquivo
print("Arquivo 'valorbase.xlsx' criado.")

#---------------------------------------------------------------------------------------------

#Filtrar produtos por preço mínimo e avaliação mínima
preco_min = float(input("\nDigite um preço mínimo: "))     #Valores do usuário
avaliacao_min = float(input("Digite uma avaliação mínima (ex: 4.0): "))

filt_aval = []  # produtos filtrados


for p in produtos_validos:        #Percorre todos os produtos

    preco = p['valor_num']      #preço

    
    padrao = re.search(r"(\d[.,]\d)", p['avaliacao'])       #Extrair a avaliação
    #Regex: \d: captura um dígito (0–9)       [.,]: captura um ponto (.) OU vírgula (,)

    if padrao:
        avaliacao = float(padrao.group().replace(",", "."))
    else:
        avaliacao = 0.0  #Se não tiver avaliação

    if preco >= preco_min and avaliacao >= avaliacao_min:       #Aplica os filtros
        filt_aval.append(p)


wbC = openpyxl.Workbook()       #Cria planilha
wsC = wbC.active
wsC.title = "preço&avaliação"


wsC.append(["Loja", "Produto", "Preço", "Avaliação", "Link"])
for cell in wsC[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')       #Cabeçalho


for p in filt_aval:       #Adiciona produto filtrado
    wsC.append([p['loja'], p['titulo'], f"R${p['valor_num']}", p['avaliacao'], p['link']])


wbC.save("preco&avaliacao.xlsx")        #Salva arquivo
print("Arquivo 'preco&avaliacao.xlsx' criado.")

#----------------------------------------------------------------------------------------------------------------
loja_escolhida = input("\nDigite a loja (Amazon ou Magalu): ").strip().lower()        #Filtra produtos pela loja


filtrados = []        #Lista pós filtros
for p in lista_produtos:
    if p['loja'].lower() == loja_escolhida:     #Compara a loja
        filtrados.append(p)


wbD = openpyxl.Workbook()       #Cria planilha
wsD = wbD.active
wsD.title = f"filtro{loja_escolhida}"
wsD.append(["Loja", "Produto", "Preço", "Avaliação", "Link"])
for cell in wsD[1]: 
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')


for p in filtrados:       #Adicionando perfils da loja
    wsD.append([p['loja'], p['titulo'], p['preco'], p['avaliacao'], p['link']])

wbD.save(f"itens{loja_escolhida}.xlsx")
print(f"Arquivo 'itens{loja_escolhida}.xlsx' criado.")

#--------------------------------------------------------------------------------------------------

#Filtra produtos pela marca do título
marca_escolhida = input("\nDigite a marca desejada: ").strip().lower()        
filtrados2 = []

for p in lista_produtos:        #Busca pela marca dentro do títul
    if marca_escolhida in p['titulo'].lower():
        filtrados2.append(p)


wbE = openpyxl.Workbook()       #Cria planilha
wsE = wbE.active
wsE.title = f"marca{marca_escolhida}"

wsE.append(["Loja", "Produto", "Preço", "Avaliação", "Link"])
for cell in wsE[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

for p in filtrados2:       #Adiciona produtos encontrados
    wsE.append([p['loja'], p['titulo'], p['preco'], p['avaliacao'], p['link']])

wbE.save(f"marca{marca_escolhida}.xlsx")
print(f"Arquivo 'marca{marca_escolhida}.xlsx' criado.")
